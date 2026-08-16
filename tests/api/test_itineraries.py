import io

from docx import Document
from openpyxl import load_workbook

from app.core.auth import create_access_token
from app.models.enums import MembershipTier
from tests.factories import create_landmark, create_member


def _headers(member) -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token(member.id)}"}


def _payload(**overrides) -> dict[str, object]:
    value: dict[str, object] = {
        "title": "高级会员行程",
        "ip_type": "game",
        "work": "黑神话：悟空",
        "origin_city": "朔州",
        "start_date": "2026-09-01",
        "end_date": "2026-09-02",
        "daily_hours": 8,
    }
    value.update(overrides)
    return value


def test_only_premium_member_can_create_and_manage_itinerary(client, db_session):
    landmark = create_landmark(db_session)
    lite = create_member(db_session, MembershipTier.LITE)
    premium = create_member(db_session, MembershipTier.PREMIUM)

    denied = client.post("/api/v1/itineraries", json=_payload(), headers=_headers(lite))
    assert denied.status_code == 403

    created = client.post("/api/v1/itineraries", json=_payload(must_visit_landmark_ids=[landmark.id]), headers=_headers(premium))
    assert created.status_code == 200
    itinerary = created.json()
    assert itinerary["days"][0]["stops"][0]["landmark_id"] == landmark.id
    assert itinerary["status"] == "succeeded"

    updated = client.patch(f"/api/v1/itineraries/{itinerary['id']}", json={"title": "已编辑行程"}, headers=_headers(premium))
    assert updated.status_code == 200
    assert updated.json()["version"] == 2
    assert updated.json()["title"] == "已编辑行程"


def test_itinerary_exports_html_docx_and_xlsx(client, db_session):
    create_landmark(db_session, landmark_name="导出地标")
    premium = create_member(db_session, MembershipTier.PREMIUM)
    created = client.post("/api/v1/itineraries", json=_payload(end_date="2026-09-01"), headers=_headers(premium))
    itinerary_id = created.json()["id"]

    html_export = client.get(f"/api/v1/itineraries/{itinerary_id}/exports/html", headers=_headers(premium))
    docx_export = client.get(f"/api/v1/itineraries/{itinerary_id}/exports/docx", headers=_headers(premium))
    xlsx_export = client.get(f"/api/v1/itineraries/{itinerary_id}/exports/xlsx", headers=_headers(premium))

    assert html_export.status_code == 200
    assert "导出地标" in html_export.content.decode("utf-8")
    assert "高级会员行程" in "\n".join(paragraph.text for paragraph in Document(io.BytesIO(docx_export.content)).paragraphs)
    workbook = load_workbook(io.BytesIO(xlsx_export.content), read_only=True)
    worksheet = workbook["行程计划"]
    assert worksheet.title == "行程计划"
    assert worksheet["A1"].value == "价格仅作参考，请以实际情况为准。"
    assert [worksheet.cell(2, column).value for column in range(1, 8)] == ["日期", "城市", "时间", "游玩地点/项目", "交通", "重点内容", "费用估算（元）"]
    assert any(row[3] == "导出地标" for row in worksheet.iter_rows(min_row=3, values_only=True))
    assert "参考说明" in workbook.sheetnames


def test_meituan_plan_is_normalized_and_budgeted_in_xlsx_export(client, db_session):
    landmark = create_landmark(db_session, landmark_name="开封府", city_name="开封市")
    premium = create_member(db_session, MembershipTier.PREMIUM)
    plan = """
| 日期 | 城市 | 时间 | 游玩地点/项目 | 交通 | 重点内容 | 费用估算(元) |
| --- | --- | --- | --- | --- | --- | --- |
| 2026-09-01 | 开封市 | 08:00-09:00 | 郑州→开封 | 高铁 | 去程 | 50元 |
|  | 开封市 | 12:00-13:00 | 午餐：灌汤包 | 步行 | 本地美食 | 40元 |
|  | 开封市 | 20:00 | 入住酒店 | 打车 | 一晚住宿 | 220元 |
"""
    response = client.post(
        "/api/v1/itineraries",
        json=_payload(
            end_date="2026-09-01",
            must_visit_landmark_ids=[landmark.id],
            confirmed_plan=True,
            lodging_mode="none",
            meituan_plan_content=plan,
        ),
        headers=_headers(premium),
    )

    assert response.status_code == 200
    itinerary = response.json()
    assert itinerary["generator_version"] == "meituan-skill-v1"
    assert itinerary["budget_summary"]["estimated_amount"] == 310
    assert itinerary["budget_summary"]["breakdown"] == {"transport": 50, "lodging": 220, "scenic": 0, "food": 40, "other": 0}
    day = itinerary["days"][0]
    assert "当日已整理 3 项安排" in day["summary"]
    assert [item["name"] for item in day["travel_context"]["meituan_items"]] == ["郑州→开封", "午餐：灌汤包", "入住酒店"]
    assert day["travel_context"]["meituan_items"][0]["time_slot"] == "08:00"
    assert day["travel_context"]["meituan_items"][0]["end_time"] == "09:00"

    export = client.get(f"/api/v1/itineraries/{itinerary['id']}/exports/xlsx", headers=_headers(premium))
    workbook = load_workbook(io.BytesIO(export.content), read_only=True, data_only=False)
    rows = list(workbook["行程计划"].iter_rows(values_only=True))
    assert any(row[3] == "郑州→开封" and row[6] == 50 for row in rows)
    assert any(row[3] == "入住酒店" and row[6] == 220 for row in rows)
    assert rows[-1][3] == "总预算"
    assert rows[0][0] == "价格仅作参考，请以实际情况为准。"
    assert rows[-1][6] == f"=SUM(G3:G{len(rows) - 1})"


def test_finalize_keeps_supplier_content_out_of_the_user_response(client, db_session, monkeypatch):
    landmark = create_landmark(db_session, landmark_name="开封府", city_name="开封市")
    premium = create_member(db_session, MembershipTier.PREMIUM)
    plan = """
| 日期 | 城市 | 时间 | 游玩地点/项目 | 交通 | 重点内容 | 费用估算(元) |
| --- | --- | --- | --- | --- | --- | --- |
| 2026-09-01 | 开封市 | 09:00 | 开封府 | 打车 | 游览 | 65元 |
"""

    from app.services.travel_enrichment import MeituanTravelPlan
    monkeypatch.setattr(
        "app.services.itinerary_planner.TravelEnrichmentService.plan",
        lambda *_: MeituanTravelPlan("available", plan, "后台来源", None),
    )
    response = client.post(
        "/api/v1/itineraries/finalize",
        json=_payload(end_date="2026-09-01", must_visit_landmark_ids=[landmark.id], lodging_mode="none"),
        headers=_headers(premium),
    )

    assert response.status_code == 200
    body = response.json()
    assert "meituan_plan" not in body
    assert body["days"][0]["travel_context"]["meituan_items"][0]["name"] == "开封府"


def test_itinerary_is_private_to_its_owner_and_can_be_deleted(client, db_session):
    create_landmark(db_session)
    owner = create_member(db_session, MembershipTier.PREMIUM)
    other = create_member(db_session, MembershipTier.PREMIUM)
    created = client.post("/api/v1/itineraries", json=_payload(end_date="2026-09-01"), headers=_headers(owner))
    itinerary_id = created.json()["id"]

    assert client.get(f"/api/v1/itineraries/{itinerary_id}", headers=_headers(other)).status_code == 404
    assert client.delete(f"/api/v1/itineraries/{itinerary_id}", headers=_headers(owner)).status_code == 204
    assert client.get(f"/api/v1/itineraries/{itinerary_id}", headers=_headers(owner)).status_code == 404
